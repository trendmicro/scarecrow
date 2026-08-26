"""XLSX hidden-column vector (structural).

Writes the payload into a cell in a column marked ``hidden="1"`` (invisible in the
Excel grid). Distinct hiding axis from hidden_row. Read by all default extractor
paths. The target cell is far off-canvas and dedicated to this vector, so it never
collides with real data; apply saves any prior value and clean restores it.
"""

from __future__ import annotations

from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

from scarecrow.core.marker import Marker
from scarecrow.vector import register
from scarecrow.formats.xlsx.vectors._common import CELLS, record

_CELL = CELLS["hidden_col"]                          # e.g. AZ1003
_COL = get_column_letter(coordinate_to_tuple(_CELL)[1])  # -> "AZ"


class HiddenColVector:
    name = "hidden_col"
    format = "xlsx"
    marker_kind = "structural"
    description = "payload in a hidden column (invisible in the grid, read by extractors)"

    def apply(self, doc: object, payload: str, marker: Marker) -> None:
        ws = doc.worksheets[0]                   # type: ignore[attr-defined]
        prev = "" if ws[_CELL].value is None else str(ws[_CELL].value)
        ws[_CELL] = payload
        ws.column_dimensions[_COL].hidden = True
        record(marker, self.name, f"{ws.title}!{_CELL}", payload, prev)

    def clean(self, doc: object, marker: Marker) -> None:
        return

    def clean_from_sidecar(self, doc: object, entry: dict) -> None:
        sheet, ref = entry["location"].split("!", 1)
        ws = doc[sheet]                          # type: ignore[index]
        ws[ref].value = (entry.get("previous") or None)
        ws.column_dimensions[_COL].hidden = False


register(HiddenColVector())
