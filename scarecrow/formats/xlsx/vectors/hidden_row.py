"""XLSX hidden-row vector (structural).

Writes the payload into a cell in a row marked ``hidden="1"`` (collapsed to zero
height in Excel). Every extractor reads the cell value; the row is invisible in
the grid unless the user notices the skipped row number and unhides it. Read by
all default paths (openpyxl, pandas, markitdown, LibreOffice CSV, Tika, raw XML).

The target cell is far off-canvas and dedicated to this vector, so it never
collides with real data; apply saves any prior value and clean restores it.
"""

from __future__ import annotations

from openpyxl.utils.cell import coordinate_to_tuple

from scarecrow.core.marker import Marker
from scarecrow.vector import register
from scarecrow.formats.xlsx.vectors._common import CELLS, record

_CELL = CELLS["hidden_row"]                 # e.g. AZ1002
_ROW = coordinate_to_tuple(_CELL)[0]        # -> 1002


class HiddenRowVector:
    name = "hidden_row"
    format = "xlsx"
    marker_kind = "structural"
    description = "payload in a hidden row (collapsed to zero height, read by extractors)"

    def apply(self, doc: object, payload: str, marker: Marker) -> None:
        ws = doc.worksheets[0]                   # type: ignore[attr-defined]
        prev = "" if ws[_CELL].value is None else str(ws[_CELL].value)
        ws[_CELL] = payload
        ws.row_dimensions[_ROW].hidden = True
        record(marker, self.name, f"{ws.title}!{_CELL}", payload, prev)

    def clean(self, doc: object, marker: Marker) -> None:
        return

    def clean_from_sidecar(self, doc: object, entry: dict) -> None:
        sheet, ref = entry["location"].split("!", 1)
        ws = doc[sheet]                          # type: ignore[index]
        ws[ref].value = (entry.get("previous") or None)
        ws.row_dimensions[_ROW].hidden = False


register(HiddenRowVector())
