"""XLSX white-on-white vector (structural).

Writes the payload into a cell with white font on a white solid fill, so the text
is invisible on the default background. Placed in a cell just outside the likely
used range so it does not overwrite real content; every extractor reads the raw
string regardless of colour.

Reversal is sidecar-driven (clear the cell we wrote). We only ever write to an
empty target cell, so clean simply blanks it.

Known limitation (dark mode / coloured backgrounds): this vector assumes a white
grid. Unlike Word and PowerPoint, where the document canvas stays white even under a
dark application theme, Excel's dark mode can darken the cell area, which would make
white-on-white text stand out. There is no clean, viewer-portable way to set a font
colour that tracks the theme background in the SpreadsheetML font model, so we do not
attempt a themeColor fix here. Prefer the flagship `very_hidden_sheet` vector, whose
concealment is background-independent, when dark-mode viewers are a concern.
"""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill

from scarecrow.core.marker import Marker
from scarecrow.vector import register
from scarecrow.formats.xlsx.vectors._common import CELLS, record

_CELL = CELLS["white_fill"]  # a dedicated off-canvas cell


class WhiteFillVector:
    name = "white_fill"
    format = "xlsx"
    marker_kind = "structural"
    description = "white text on white fill (invisible on the default background)"

    def apply(self, doc: object, payload: str, marker: Marker) -> None:
        ws = doc.worksheets[0]                   # type: ignore[attr-defined]
        cell = ws[_CELL]
        prev = "" if cell.value is None else str(cell.value)
        cell.value = payload
        cell.font = Font(color="FFFFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
        record(marker, self.name, f"{ws.title}!{_CELL}", payload, prev)

    def clean(self, doc: object, marker: Marker) -> None:
        return

    def clean_from_sidecar(self, doc: object, entry: dict) -> None:
        sheet, ref = entry["location"].split("!", 1)
        ws = doc[sheet]                          # type: ignore[index]
        ws[ref].value = (entry.get("previous") or None)
        ws[ref].font = Font()
        ws[ref].fill = PatternFill()


register(WhiteFillVector())
