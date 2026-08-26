"""Layer-1 tests for the XLSX vector library (Slice 6).

Deterministic, offline. Mirrors the DOCX/PPTX vector tests:
  - all 10 XLSX vectors are registered, in order
  - each vector applied alone: its payload reaches the file, and clean removes it
    (visible data preserved; sidecar restored)
  - --all applies the default (non-research-only) set and clean reverses it
  - the two encoding vectors are excluded from --all but stay in the library
  - real data in the target cells is never clobbered (off-canvas cells + prev-value)
  - inspect reports vector + location; sidecar always written (XLSX is sidecar-only)
"""
from __future__ import annotations

import zipfile

import pytest
from openpyxl import Workbook, load_workbook

from scarecrow.formats import xlsx as _xlsx  # noqa: F401  (registers vectors)
from scarecrow import vector
from scarecrow.core.marker import Marker, write_sidecar, sidecar_path_for
from scarecrow.core.modes import Mode
from scarecrow.core.payloads import select_payload
from scarecrow.formats.xlsx.reverse import inspect_xlsx, clean_xlsx

ALL = ["very_hidden_sheet", "semicolon_format", "white_fill", "hidden_row",
       "hidden_col", "header_footer", "comment", "core_props"]
# XLSX has no in-text encoding vectors: openpyxl strips those codepoints on save
# (and they are ineffective anyway). See vectors/__init__.py.
PAYLOAD_MARK = "protected against automated ai"
# Real content in the "used range" — the target cells are off-canvas (AZ1000+),
# so protect/clean must never touch these.
VISIBLE = {"A1": "Metric", "B1": "Value", "A2": "Q3 revenue", "B2": 4200000}


def _make_original(tmp_path):
    p = tmp_path / "book.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Financials"
    for ref, val in VISIBLE.items():
        ws[ref] = val
    wb.save(str(p))
    return p


def _protect(orig, tmp_path, vectors: list[str]):
    out = tmp_path / "book.scarecrow.xlsx"
    wb = load_workbook(str(orig))
    marker = Marker()
    payload = select_payload(Mode.REFUSE, None)
    for name in vectors:
        vector.get(name, "xlsx").apply(wb, payload, marker)
    wb.save(str(out))
    write_sidecar(str(out), marker.sidecar_entries)
    return out


def _all_xml(path) -> str:
    z = zipfile.ZipFile(str(path))
    return "".join(z.read(n).decode("utf-8", "replace")
                   for n in z.namelist() if n.endswith(".xml"))


def _visible_intact(path) -> bool:
    ws = load_workbook(str(path))["Financials"]
    return all(ws[ref].value == val for ref, val in VISIBLE.items())


def test_all_xlsx_vectors_registered():
    names = [v.name for v in vector.all_vectors("xlsx")]
    assert names == ALL  # order matters for --all / list-vectors


@pytest.mark.parametrize("name", ALL)
def test_vector_payload_reaches_the_file(name, tmp_path):
    orig = _make_original(tmp_path)
    prot = _protect(orig, tmp_path, [name])
    xml = _all_xml(prot).lower()
    assert (PAYLOAD_MARK in xml) or ("owner-protected" in xml)


@pytest.mark.parametrize("name", ALL)
def test_vector_clean_restores_original(name, tmp_path):
    orig = _make_original(tmp_path)
    prot = _protect(orig, tmp_path, [name])
    cleaned = tmp_path / "cleaned.xlsx"
    clean_xlsx(str(prot), str(cleaned))
    assert _visible_intact(cleaned)              # real data preserved
    xml = _all_xml(cleaned).lower()
    assert PAYLOAD_MARK not in xml
    assert "owner-protected" not in xml


def test_all_applies_every_vector_and_clean_reverses(tmp_path):
    # XLSX has no research-only vectors, so --all == the full library
    orig = _make_original(tmp_path)
    defaults = [v.name for v in vector.default_vectors("xlsx")]
    assert defaults == ALL
    prot = _protect(orig, tmp_path, defaults)
    found = {i.vector for i in inspect_xlsx(str(prot))}
    assert found == set(defaults)
    cleaned = tmp_path / "cleaned.xlsx"
    removed = clean_xlsx(str(prot), str(cleaned))
    assert removed == len(defaults)
    assert inspect_xlsx(str(cleaned)) == []
    assert _visible_intact(cleaned)


def test_very_hidden_sheet_state_and_removal(tmp_path):
    orig = _make_original(tmp_path)
    prot = _protect(orig, tmp_path, ["very_hidden_sheet"])
    wb = load_workbook(str(prot))
    vh = [s for s in wb.worksheets if s.sheet_state == "veryHidden"]
    assert len(vh) == 1 and PAYLOAD_MARK in (vh[0]["A1"].value or "").lower()
    cleaned = tmp_path / "cleaned.xlsx"
    clean_xlsx(str(prot), str(cleaned))
    assert [s.title for s in load_workbook(str(cleaned)).worksheets] == ["Financials"]


def test_real_data_never_clobbered(tmp_path):
    # even with --all, the used-range cells are untouched (payloads go off-canvas)
    orig = _make_original(tmp_path)
    defaults = [v.name for v in vector.default_vectors("xlsx")]
    prot = _protect(orig, tmp_path, defaults)
    assert _visible_intact(prot)                 # intact right after protect
    cleaned = tmp_path / "cleaned.xlsx"
    clean_xlsx(str(prot), str(cleaned))
    assert _visible_intact(cleaned)


def test_sidecar_always_written(tmp_path):
    # XLSX is sidecar-only: any protect run writes a sidecar
    orig = _make_original(tmp_path)
    prot = _protect(orig, tmp_path, ["hidden_row"])
    assert sidecar_path_for(str(prot)).exists()


def test_inspect_reports_vector_and_location(tmp_path):
    orig = _make_original(tmp_path)
    prot = _protect(orig, tmp_path, ["very_hidden_sheet", "core_props"])
    found = {i.vector: i.where for i in inspect_xlsx(str(prot))}
    assert found["very_hidden_sheet"].startswith("sidecar:sheet:")
    assert found["core_props"] == "sidecar:subject"
